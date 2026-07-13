from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable

import pandas as pd
from openpyxl import load_workbook


CDR_IGNORED_SHEETS = {
    'MASTER',
    'KPI Definition',
    'KPI Definition_OLD',
    'RANKING',
    'RANKING OnDemand',
    'RANKING 5G',
    'Lists',
    'SetupEnviroment',
}
CAMPAIGN_PATTERN = re.compile(r'(?P<market>[A-Z]{2})_Q(?P<quarter>\d)_(?P<year>\d{4})')


@dataclass(slots=True)
class DatasetSummary:
    rows: int
    columns: list[str]
    numeric_columns: list[str]
    categorical_columns: list[str]


def _first_available_series(df: pd.DataFrame, candidates: Iterable[str], default: object = pd.NA) -> pd.Series:
    for column in candidates:
        if column in df.columns:
            return df[column]
    return pd.Series([default] * len(df), index=df.index, dtype='object')


def _parse_campaign_dimension(value: object, part: str) -> object:
    if pd.isna(value):
        return pd.NA
    match = CAMPAIGN_PATTERN.search(str(value))
    if not match:
        return pd.NA
    if part == 'market':
        return match.group('market')
    if part == 'year':
        return int(match.group('year'))
    if part == 'quarter':
        return f"Q{match.group('quarter')}"
    if part == 'period':
        return f"{match.group('year')}-Q{match.group('quarter')}"
    return pd.NA


def _count_items(value: object) -> float:
    if pd.isna(value):
        return 0.0
    text = str(value).strip()
    if not text:
        return 0.0
    if '->' in text:
        return float(text.count('->') + 1)
    if '],' in text:
        return float(text.count('],') + 1)
    if ',' in text and '[' in text:
        return float(text.count(',') + 1)
    return 1.0


def _average_columns(df: pd.DataFrame, candidates: Iterable[str]) -> pd.Series:
    numeric_columns = [column for column in candidates if column in df.columns]
    if not numeric_columns:
        return pd.Series([pd.NA] * len(df), index=df.index, dtype='Float64')
    numeric_frame = df[numeric_columns].apply(pd.to_numeric, errors='coerce')
    return numeric_frame.mean(axis=1).astype('Float64')


def _sheet_has_only_empty_values(values: tuple[object, ...]) -> bool:
    return all(value is None or str(value).strip() == '' for value in values)


def _read_openxml_sheet(worksheet, progress_callback: Callable[[int], None] | None, progress_state: dict[str, int], total_rows: int) -> pd.DataFrame:
    rows_iter = worksheet.iter_rows(values_only=True)
    header: list[str] | None = None
    records: list[dict[str, object]] = []

    def advance_progress() -> None:
        progress_state['processed_rows'] += 1
        if not progress_callback or total_rows <= 0:
            return
        progress = min(55, 14 + int((progress_state['processed_rows'] / total_rows) * 41))
        if progress > progress_state['last_progress']:
            progress_state['last_progress'] = progress
            progress_callback(progress)

    for row in rows_iter:
        advance_progress()
        values = tuple(row or ())
        if _sheet_has_only_empty_values(values):
            continue
        header = [str(value).strip() if value is not None else '' for value in values]
        break

    if not header:
        return pd.DataFrame()

    for row in rows_iter:
        advance_progress()
        values = tuple(row or ())
        if _sheet_has_only_empty_values(values):
            continue
        padded = list(values[:len(header)])
        if len(padded) < len(header):
            padded.extend([None] * (len(header) - len(padded)))
        records.append(dict(zip(header, padded, strict=False)))

    if not records:
        return pd.DataFrame(columns=header)
    return pd.DataFrame.from_records(records, columns=header)


def infer_dataset_kind(df: pd.DataFrame, file_name: str = '') -> str:
    lower_name = file_name.lower()
    normalized_name = lower_name.replace('_', ' ').replace('-', ' ')
    if 'speech' in normalized_name:
        return 'speech'
    if 'voice' in normalized_name:
        return 'voice'
    if 'data' in normalized_name:
        return 'data'
    columns = set(df.columns)
    voice_markers = {
        'POLQA_LQ_Avg', 'Call_Setup_Time', 'Call_Duration', 'Call_Status',
        'Dropped_in_first_70s', 'Disturbed_Call', 'Impaired_Call',
    }
    speech_markers = {
        'LQ', 'RTP_Jitter_Avg_A', 'RTP_Jitter_Avg_B', 'RTP_Packet_Loss_A',
        'RTP_Packet_Loss_B', 'Receive_Delay', 'Listening_Technology',
        'Recording_Technology', 'Playing_Handovers',
    }
    data_markers = {
        'Mean_Data_Rate', 'TCP_Throughput', 'Data_Throughput', 'Test_Result',
        'Transfer_Access_Duration', 'Transfer_Duration',
        'TCP_RTT_Service_Access_Delay', 'DNS_Service_Access_Delay',
        'VideoStream_Time_To_Start_Buffering', 'VideoStream_Video_Stream_Duration',
        'http_Browser_Access_Duration',
    }
    scores = {
        'voice': len(columns & voice_markers),
        'speech': len(columns & speech_markers),
        'data': len(columns & data_markers),
    }
    best_kind = max(scores, key=scores.get)
    if scores[best_kind] > 0:
        return best_kind
    return 'generic'


def _normalise_dataset(df: pd.DataFrame, file_path: Path) -> pd.DataFrame:
    dataset = df.copy()
    dataset_kind = infer_dataset_kind(dataset, file_path.name)
    dataset['dataset_kind'] = dataset_kind
    dataset['source_file'] = file_path.name

    dataset['campaign'] = _first_available_series(dataset, ['Campaign'])
    derived_market = dataset['campaign'].map(lambda value: _parse_campaign_dimension(value, 'market'))
    derived_period = dataset['campaign'].map(lambda value: _parse_campaign_dimension(value, 'period'))
    existing_market = _first_available_series(dataset, ['market', 'Market'])
    existing_period = _first_available_series(dataset, ['period', 'Period'])
    dataset['market'] = existing_market.where(existing_market.notna(), derived_market)
    dataset['period'] = existing_period.where(existing_period.notna(), derived_period)
    dataset['campaign_year'] = dataset['campaign'].map(lambda value: _parse_campaign_dimension(value, 'year'))
    dataset['campaign_quarter'] = dataset['campaign'].map(lambda value: _parse_campaign_dimension(value, 'quarter'))

    dataset['operator'] = _first_available_series(dataset, ['Operator', 'Home_Operator', 'Home_Operator_A'])
    dataset['session_type'] = _first_available_series(dataset, ['Session_Type', 'Type_of_Test'])
    dataset['test_name'] = _first_available_series(dataset, ['Test_Name', 'Session_Type'])
    dataset['direction'] = _first_available_series(dataset, ['Direction', 'Call_Direction'])
    dataset['region'] = _first_available_series(dataset, ['Region'])
    dataset['vendor'] = _first_available_series(dataset, ['Vendor'])
    dataset['status'] = _first_available_series(dataset, ['Call_Status', 'Test_Result', 'Test_Status'])

    dataset['disturbed'] = _first_available_series(dataset, ['Disturbed_Call']).astype(str).str.lower().eq('yes')
    dataset['impaired'] = _first_available_series(dataset, ['Impaired_Call']).astype(str).str.lower().eq('yes')
    dataset['dropped'] = dataset['status'].astype(str).str.contains('drop', case=False, na=False)
    if 'Dropped_in_first_70s' in dataset.columns:
        dataset['dropped'] = dataset['dropped'] | dataset['Dropped_in_first_70s'].astype(str).str.lower().eq('yes')
    if 'Unsustainable_Call' in dataset.columns:
        dataset['unsustainable_call'] = dataset['Unsustainable_Call'].astype(str).str.lower().eq('yes')
    else:
        dataset['unsustainable_call'] = False

    status_normalized = dataset['status'].astype(str).str.strip().str.lower()
    dataset['success'] = status_normalized.isin({'completed', 'success', 'ok', 'passed'})
    dataset['failure'] = dataset['status'].notna() & ~dataset['success']

    dataset['event_start_time'] = pd.to_datetime(
        _first_available_series(dataset, ['Call_Start_Time', 'Test_Start_Time', 'Data_Start_Time']),
        errors='coerce',
    )
    dataset['event_end_time'] = pd.to_datetime(
        _first_available_series(dataset, ['Call_End_Time', 'Test_End_Time', 'Data_End_Time']),
        errors='coerce',
    )
    dataset['hour_bucket'] = dataset['event_start_time'].dt.hour
    dataset['day_bucket'] = dataset['event_start_time'].dt.day

    dataset['setup_time_seconds'] = pd.to_numeric(
        _first_available_series(
            dataset,
            ['Call_Setup_Time', 'Transfer_Access_Duration', 'http_Browser_Access_Duration', 'VideoStream_Time_To_Start_Buffering'],
        ),
        errors='coerce',
    )
    dataset['duration_seconds'] = pd.to_numeric(
        _first_available_series(
            dataset,
            ['Call_Duration', 'Test_Duration', 'Data_Test_Duration', 'Transfer_Duration', 'VideoStream_Video_Stream_Duration'],
        ),
        errors='coerce',
    )
    dataset['quality_score'] = pd.to_numeric(
        _first_available_series(dataset, ['POLQA_LQ_Avg', 'LQ', 'Mean_Data_Rate']),
        errors='coerce',
    )
    dataset['throughput_mbps'] = pd.to_numeric(
        _first_available_series(dataset, ['Mean_Data_Rate', 'TCP_Throughput', 'Data_Throughput']),
        errors='coerce',
    )
    dataset['latency_ms'] = pd.to_numeric(
        _first_available_series(dataset, ['Receive_Delay', 'TCP_RTT_Service_Access_Delay', 'DNS_Service_Access_Delay']),
        errors='coerce',
    )
    dataset['packet_loss_pct'] = _average_columns(dataset, ['RTP_Packet_Loss_A', 'RTP_Packet_Loss_B', 'Packet_Loss_Score'])
    dataset['jitter_ms'] = _average_columns(dataset, ['RTP_Jitter_Avg_A', 'RTP_Jitter_Avg_B'])
    dataset['handovers'] = _first_available_series(dataset, ['Handovers_Info', 'Handovers_Info_A', 'Playing_Handovers']).map(_count_items)

    dataset['technology_primary'] = _first_available_series(
        dataset,
        ['L2_call_Mode_A', 'RAT_A', 'Playing_Technology', 'PCell_RAT_Timeline', 'RAT'],
    )
    dataset['technology_secondary'] = _first_available_series(
        dataset,
        ['L2_call_Mode_B', 'RAT_B', 'Recording_Technology', 'RAT_Timeline'],
    )

    return dataset


def _load_excel_dataset(file_path: Path, progress_callback: Callable[[int], None] | None = None) -> pd.DataFrame:
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    candidate_sheets = [sheet_name for sheet_name in workbook.sheetnames if sheet_name not in CDR_IGNORED_SHEETS]
    total_rows = sum(max(workbook[sheet_name].max_row or 0, 1) for sheet_name in candidate_sheets) or 1

    if progress_callback:
        progress_callback(14)

    data_frames: list[pd.DataFrame] = []
    progress_state = {'processed_rows': 0, 'last_progress': 14}
    for sheet_name in candidate_sheets:
        sheet = _read_openxml_sheet(workbook[sheet_name], progress_callback, progress_state, total_rows)
        sheet = sheet.dropna(axis=0, how='all').dropna(axis=1, how='all')
        if sheet.empty:
            continue
        if all(str(column).startswith('Unnamed:') for column in sheet.columns):
            continue
        sheet['source_sheet'] = sheet_name
        data_frames.append(sheet)

    if not data_frames:
        df = pd.read_excel(file_path)
        if progress_callback:
            progress_callback(55)
        return df

    if len(data_frames) == 1 and data_frames[0]['source_sheet'].nunique() == 1 and data_frames[0]['source_sheet'].iat[0] not in {
        'Vodafone',
        'Telefonica',
        'Operator4',
        'Operator5',
    }:
        dataset = _normalise_dataset(data_frames[0], file_path)
        if progress_callback:
            progress_callback(55)
        return dataset

    combined = pd.concat(data_frames, ignore_index=True, sort=False)
    dataset = _normalise_dataset(combined, file_path)
    if progress_callback:
        progress_callback(55)
    return dataset


def load_dataset(file_path: Path, progress_callback: Callable[[int], None] | None = None) -> pd.DataFrame:
    suffix = file_path.suffix.lower()
    if suffix == '.csv':
        df = _normalise_dataset(pd.read_csv(file_path), file_path)
        if progress_callback:
            progress_callback(55)
        return df
    if suffix in {'.xlsx', '.xlsm'}:
        return _load_excel_dataset(file_path, progress_callback=progress_callback)
    if suffix == '.xls':
        df = _normalise_dataset(pd.read_excel(file_path), file_path)
        if progress_callback:
            progress_callback(55)
        return df
    raise ValueError(f'Unsupported file type: {suffix}')


def summarise_dataset(df: pd.DataFrame) -> DatasetSummary:
    numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
    categorical_columns = [column for column in df.columns.tolist() if column not in numeric_columns]
    return DatasetSummary(
        rows=len(df.index),
        columns=df.columns.tolist(),
        numeric_columns=numeric_columns,
        categorical_columns=categorical_columns,
    )
